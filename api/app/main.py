import os
import io
import re
import sqlite3
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

# ----- Optional S3 (R2 / Backblaze / Supabase-compatible) -----
import boto3
from botocore.client import Config
from botocore.exceptions import ClientError

# ----- Excel parsing (for /timesheet/parse) -----
from openpyxl import load_workbook
import xlrd


# =============================================================================
# Config
# =============================================================================
DATA_DIR = os.getenv("DATA_DIR", "/data")
os.makedirs(DATA_DIR, exist_ok=True)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Use your v4 DB (you can change this to v3/v5 etc. when you need)
DB_PATH = os.path.join(BASE_DIR, "employees_with_company_v4.db")

# S3-compatible object storage (optional)
S3_ENDPOINT = (os.getenv("S3_ENDPOINT") or "").strip()
S3_REGION = os.getenv("S3_REGION", "us-east-1")
S3_ACCESS_KEY_ID = os.getenv("S3_ACCESS_KEY_ID")
S3_SECRET_ACCESS_KEY = os.getenv("S3_SECRET_ACCESS_KEY")
S3_BUCKET = os.getenv("S3_BUCKET")

def _normalize_endpoint(ep: str) -> str:
    ep = ep.strip()
    if ep.startswith("https://"):
        ep = ep[len("https://"):]
    if ep.startswith("http://"):
        ep = ep[len("http://"):]
    return ep

S3_HOST = _normalize_endpoint(S3_ENDPOINT) if S3_ENDPOINT else ""


# =============================================================================
# App
# =============================================================================
app = FastAPI(title="YHO API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],          # tighten in prod
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    max_age=600,
)


# =============================================================================
# DB helpers
# =============================================================================
def open_db() -> sqlite3.Connection:
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(f"DB not found at {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def list_tables(conn: sqlite3.Connection) -> List[str]:
    t = conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
    return [r["name"] for r in t]

def dict_from_row(r: sqlite3.Row) -> Dict[str, Any]:
    return {k: r[k] for k in r.keys()}

def first_table_with_column(conn: sqlite3.Connection, colname: str) -> Optional[Tuple[str, List[str]]]:
    for t in list_tables(conn):
        cols = [c["name"] for c in conn.execute(f"PRAGMA table_info('{t}')").fetchall()]
        if colname in cols:
            return t, cols
    return None

def discover_employee_table(conn: sqlite3.Connection) -> Tuple[str, List[str]]:
    """
    Prefer a table literally named 'employees', else any table with 'employee_id' column,
    else the first table found.
    """
    tables = list_tables(conn)
    if not tables:
        raise RuntimeError("No tables found in SQLite DB.")

    # Prefer exact 'employees'
    for t in tables:
        if t.lower() == "employees":
            cols = [c["name"] for c in conn.execute(f"PRAGMA table_info('{t}')").fetchall()]
            return t, cols

    # Else table with employee_id
    found = first_table_with_column(conn, "employee_id")
    if found:
        return found

    # Else first table
    tname = tables[0]
    cols = [c["name"] for c in conn.execute(f"PRAGMA table_info('{tname}')").fetchall()]
    return tname, cols

def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s:
            return None
        s = s.replace("$", "").replace(",", "")
        return float(s)
    except Exception:
        return None

def compute_labor_rate_display(row: Dict[str, Any]) -> Optional[float]:
    for key in ("labor_rate", "pay_rate", "payrate", "rate", "hourly_rate"):
        if key in row:
            n = to_float(row.get(key))
            if n is not None:
                return n
    return None

def filter_payload_to_table(payload: Dict[str, Any], columns: List[str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k, v in payload.items():
        if k in columns:
            if k in ("labor_rate", "per_diem", "pay_rate", "payrate", "hourly_rate"):
                out[k] = to_float(v)
            else:
                out[k] = v
    return out


# =============================================================================
# Employees
# =============================================================================
@app.get("/employees")
def list_employees(limit: int = Query(1000, ge=1, le=10000), offset: int = Query(0, ge=0)):
    try:
        with open_db() as conn:
            tname, cols = discover_employee_table(conn)
            total = conn.execute(f"SELECT COUNT(*) AS c FROM '{tname}'").fetchone()["c"]
            rs = conn.execute(f"SELECT * FROM '{tname}' LIMIT ? OFFSET ?", (limit, offset)).fetchall()
            rows = [dict_from_row(r) for r in rs]
            for row in rows:
                row["labor_rate_display"] = compute_labor_rate_display(row)
            return {"rows": rows, "total": total, "limit": limit, "offset": offset, "table": tname, "columns": cols}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/employees/{employee_id}")
def get_employee(employee_id: str):
    try:
        with open_db() as conn:
            tname, cols = discover_employee_table(conn)
            r = conn.execute(f"SELECT * FROM '{tname}' WHERE employee_id = ?", (employee_id,)).fetchone()
            if not r:
                raise HTTPException(status_code=404, detail="Not found")
            row = dict_from_row(r)
            row["labor_rate_display"] = compute_labor_rate_display(row)
            return row
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/employees")
def create_employee(payload: Dict[str, Any]):
    if not payload or not payload.get("employee_id"):
        raise HTTPException(status_code=400, detail="employee_id is required")
    try:
        with open_db() as conn:
            tname, cols = discover_employee_table(conn)
            data = filter_payload_to_table(payload, cols)
            if not data:
                raise HTTPException(status_code=400, detail="No valid fields to insert")
            keys = list(data.keys())
            qmarks = ",".join(["?"] * len(keys))
            sql = f"INSERT INTO '{tname}' ({','.join(keys)}) VALUES ({qmarks})"
            conn.execute(sql, tuple(data[k] for k in keys))
            conn.commit()
            return {"ok": True}
    except sqlite3.IntegrityError as e:
        raise HTTPException(status_code=409, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/employees/{employee_id}")
def patch_employee(employee_id: str, payload: Dict[str, Any]):
    if not payload:
        return {"ok": True, "updated": 0}
    try:
        with open_db() as conn:
            tname, cols = discover_employee_table(conn)
            data = filter_payload_to_table(payload, cols)
            if not data:
                return {"ok": True, "updated": 0}
            sets = ", ".join([f"{k}=?" for k in data.keys()])
            sql = f"UPDATE '{tname}' SET {sets} WHERE employee_id=?"
            cur = conn.execute(sql, (*[data[k] for k in data.keys()], employee_id))
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Employee not found")
            return {"ok": True, "updated": cur.rowcount}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Documents (S3-compatible). If not configured, returns empty list for /documents.
# =============================================================================
def s3_client():
    if not (S3_HOST and S3_ACCESS_KEY_ID and S3_SECRET_ACCESS_KEY and S3_BUCKET):
        raise RuntimeError("S3 not fully configured; set S3_ENDPOINT, S3_ACCESS_KEY_ID, S3_SECRET_ACCESS_KEY, S3_BUCKET.")
    return boto3.client(
        "s3",
        region_name=S3_REGION,
        aws_access_key_id=S3_ACCESS_KEY_ID,
        aws_secret_access_key=S3_SECRET_ACCESS_KEY,
        endpoint_url=f"https://{S3_HOST}",
        config=Config(s3={"addressing_style": "path"}),
    )

@app.get("/documents")
def list_documents(limit: int = 500):
    try:
        s3 = s3_client()
        paginator = s3.get_paginator("list_objects_v2")
        pages = paginator.paginate(Bucket=S3_BUCKET, PaginationConfig={"MaxItems": limit})
        rows: List[Dict[str, Any]] = []
        for page in pages:
            for obj in page.get("Contents", []) or []:
                rows.append({
                    "key": obj["Key"],
                    "size": obj.get("Size", 0),
                    "last_modified": obj.get("LastModified").isoformat() if obj.get("LastModified") else None,
                })
                if len(rows) >= limit:
                    break
            if len(rows) >= limit:
                break
        return {"rows": rows, "total": len(rows)}
    except RuntimeError as e:
        # not configured -> don't break the whole app
        return {"rows": [], "total": 0, "note": str(e)}
    except ClientError as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/documents/upload")
async def upload_document(
    employee_name: str = Form(""),
    doc_types: str = Form(""),
    custom_type: str = Form(""),
    file: UploadFile = File(...),
):
    try:
        s3 = s3_client()
        name_key = "_".join((employee_name or "unknown").split()) or "unknown"
        folder = (custom_type or doc_types or "misc").strip() or "misc"
        filename = file.filename or "upload.bin"
        key = f"{name_key}/{folder}/{filename}"

        body = await file.read()
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=key,
            Body=body,
            ContentType=file.content_type or "application/octet-stream",
        )
        return {"ok": True, "key": key}
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except ClientError as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/documents/{key:path}/download")
def download_document(key: str):
    try:
        s3 = s3_client()
        obj = s3.get_object(Bucket=S3_BUCKET, Key=key)
        content_type = obj.get("ContentType") or "application/octet-stream"
        data = obj["Body"].read()
        return StreamingResponse(
            io.BytesIO(data),
            media_type=content_type,
            headers={"Content-Disposition": f'inline; filename="{os.path.basename(key)}"'},
        )
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except ClientError as e:
        code = e.response.get("ResponseMetadata", {}).get("HTTPStatusCode", 500)
        raise HTTPException(status_code=code, detail=str(e))


# =============================================================================
# Timesheet parsing -> matched/unmatched (Week 1 hours fill)
# =============================================================================

def _norm_name(s: str) -> str:
    if s is None:
        return ""
    s = re.sub(r"\(\s*\d+\s*\)", "", s)  # remove "(123)"
    s = re.sub(r"\s+", " ", s).strip()
    return s.upper()

def _as_hours(v):
    if v is None:
        return None
    try:
        s = str(v).strip()
        if not s:
            return None
        s = s.replace(",", "")
        return float(s)
    except Exception:
        return None

def _other_nonempty_in_row(cells, skip_idx: int):
    """
    Given a list of cell values and an index to skip (the known label cell),
    return:
      - index of the only other non-empty cell (if exactly one exists)
      - and its value
    otherwise (0 or >1 others) return (None, None).
    """
    idxs = [i for i, v in enumerate(cells) if (v not in (None, "") and i != skip_idx)]
    if len(idxs) == 1:
        return idxs[0], cells[idxs[0]]
    return None, None

def _process_rows(rows_2d, out_pairs):
    """
    rows_2d: list[list[Any]] for a worksheet
    out_pairs: list to append dicts like {"name": <str>, "hours": <float>}
    Implements the rule:
      Row with 'Employee' -> the only other non-empty cell is the NAME
      Next row with 'Total Hours' -> the only other non-empty cell is HOURS
    """
    i = 0
    n = len(rows_2d)
    while i < n:
        row = rows_2d[i]
        # Find 'Employee' in current row (case-insensitive)
        emp_idx = None
        for j, v in enumerate(row):
            if isinstance(v, str) and v.strip().lower() == "employee":
                emp_idx = j
                break
        if emp_idx is None:
            i += 1
            continue

        # Extract the single name cell in this row
        name_col, name_val = _other_nonempty_in_row(row, emp_idx)
        if name_col is None or not isinstance(name_val, str) or not name_val.strip():
            # malformed "Employee" row; skip forward
            i += 1
            continue
        name_text = name_val.strip()

        # Advance to find the next row that contains "Total Hours"
        k = i + 1
        hours_val = None
        while k < n:
            r2 = rows_2d[k]
            th_idx = None
            for jj, vv in enumerate(r2):
                if isinstance(vv, str) and vv.strip().lower() == "total hours":
                    th_idx = jj
                    break
            if th_idx is not None:
                # Extract the single hours cell in this row
                hrs_col, hrs_raw = _other_nonempty_in_row(r2, th_idx)
                hours_val = _as_hours(hrs_raw) if hrs_col is not None else None
                break
            k += 1

        # Record pair if we found hours; otherwise still record with hours=None
        out_pairs.append({"name": name_text, "hours": hours_val})
        # Continue scanning AFTER the Total Hours row if found, else just move on
        i = (k + 1) if (k < n) else (i + 1)

@app.post("/timesheet/parse")
async def parse_timesheet(file: UploadFile = File(...)):
    """
    Supports .xlsx/.xlsm (openpyxl) and .xls (xlrd).
    Pattern:
      <Row>: 'Employee' + (only other non-empty) -> NAME
      <Next Row with 'Total Hours'>: 'Total Hours' + (only other non-empty) -> HOURS
    Output:
      { matched: [{employee_id, name, hours}], unmatched: [{name, reason, hours}] }
    """
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty upload")

    # Detect format
    is_zip = raw[:4] == b"PK\x03\x04"          # xlsx/xlsm
    is_ole = raw[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"  # xls

    wb_xlsx = None
    wb_xls = None
    try:
        if is_zip:
            wb_xlsx = load_workbook(io.BytesIO(raw), data_only=True)
        elif is_ole:
            wb_xls = xlrd.open_workbook(file_contents=raw)
        else:
            # try xlsx then xls
            try:
                wb_xlsx = load_workbook(io.BytesIO(raw), data_only=True)
            except Exception:
                wb_xls = xlrd.open_workbook(file_contents=raw)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {e}")

    # Pull employees
    try:
        with open_db() as conn:
            tname, _ = discover_employee_table(conn)
            rows_emp = conn.execute(f"SELECT * FROM '{tname}'").fetchall()
            emps = [dict_from_row(r) for r in rows_emp]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB error: {e}")

    by_ts = {}
    by_nm = {}
    for r in emps:
        ts = _norm_name(str(r.get("timesheet_name") or ""))
        nm = _norm_name(str(r.get("name") or ""))
        if ts:
            by_ts[ts] = r
        if nm and nm not in by_nm:
            by_nm[nm] = r

    # Parse workbook into (name, hours) pairs using the new rule
    pairs = []

    if wb_xlsx is not None:
        for ws in wb_xlsx.worksheets:
            rows = []
            # read row values (keep empty cells so "only other cell" logic works)
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            for rr in range(1, max_row + 1):
                rows.append([ws.cell(rr, cc).value for cc in range(1, max_col + 1)])
            _process_rows(rows, pairs)

    if wb_xls is not None:
        for sh in wb_xls.sheets():
            rows = []
            for rr in range(sh.nrows):
                rows.append([sh.cell(rr, cc).value for cc in range(sh.ncols)])
            _process_rows(rows, pairs)

    # Match pairs to employees
    matched = []
    unmatched = []
    for p in pairs:
        raw_name = p["name"]
        hrs = p["hours"]
        key = _norm_name(raw_name)
        emp = by_ts.get(key) or by_nm.get(key)
        if emp:
            matched.append({
                "employee_id": str(emp.get("employee_id") or "").strip(),
                "name": emp.get("name") or emp.get("timesheet_name") or raw_name,
                "hours": float(hrs or 0.0),
            })
        else:
            unmatched.append({
                "name": raw_name,
                "reason": "No matching employee by timesheet_name or name",
                "hours": hrs,
            })

    # Deduplicate matched by employee_id, keep largest hours
    best = {}
    for m in matched:
        eid = m["employee_id"]
        if not eid:
            continue
        prev = best.get(eid)
        if (prev is None) or (m["hours"] or 0) > (prev["hours"] or 0):
            best[eid] = m

    return {"matched": list(best.values()), "unmatched": unmatched}
# ---- end drop-in ------------------------------------------------------------

# =============================================================================
# Debug / Health
# =============================================================================
@app.get("/debug/health")
def health():
    s3_ok = bool(S3_HOST and S3_ACCESS_KEY_ID and S3_SECRET_ACCESS_KEY and S3_BUCKET)
    return {
        "ok": True,
        "db_path": DB_PATH,
        "db_exists": os.path.exists(DB_PATH),
        "s3_configured": s3_ok,
        "bucket": S3_BUCKET if s3_ok else None,
    }

@app.get("/debug/db")
def debug_db():
    with open_db() as conn:
        info: Dict[str, Any] = {"tables": [], "meta": []}
        tables = list_tables(conn)
        info["tables"] = tables
        for t in tables:
            cols = [c["name"] for c in conn.execute(f"PRAGMA table_info('{t}')").fetchall()]
            sample = conn.execute(f"SELECT * FROM '{t}' LIMIT 1").fetchone()
            sample_dict = dict_from_row(sample) if sample else None
            info["meta"].append({"table": t, "columns": cols, "sample": sample_dict})
        return info

@app.get("/debug/sample")
def debug_sample(limit: int = 3):
    with open_db() as conn:
        tables = list_tables(conn)
        if not tables:
            return {"tables": [], "rows": []}
        tname = next((t for t in tables if t.lower() == "employees"), tables[0])
        rs = conn.execute(f"SELECT * FROM '{tname}' LIMIT ?", (limit,)).fetchall()
        rows = [dict_from_row(r) for r in rs]
        for row in rows:
            row["labor_rate_display"] = compute_labor_rate_display(row)
        return {"table": tname, "rows": rows}


